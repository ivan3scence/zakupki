from hashlib import new
from bs4 import BeautifulSoup as bs
from numpy import record
import requests
import re
import shutil
import openpyxl
from requests.api import post
import pandas as pd
import datetime
from mail import send_email

id_region = '5277335'                                                                                # id региона
thr = 5                                                                                              # кол-во потоков
to = '.'.join(datetime.date.today().strftime("%d/%m/%Y").split('/'))                                 #
frm = '.'.join((datetime.date.today() - datetime.timedelta(days=3)).strftime("%d/%m/%Y").split('/')) # 
record_file = f'zakupki_{to.split(".")[0]}_{to.split(".")[1]}_{to.split(".")[2]}.xlsx'               # сегодняшний файл
base_file = "zakupki_rastorjenie_base copy.xlsx"                                                          # база
 
 
pages = 1

def urrll(i):
    url = f'https://zakupki.gov.ru/epz/dizk/search/results.html?morphology=on&search-filter=%D0%94%D0%B0%D1%82%D0%B5+%D1%80%D0%B0%D0%B7%D0%BC%D0%B5%D1%89%D0%B5%D0%BD%D0%B8%D1%8F&sortBy=UPDATE_DATE&pageNumber=1&sortDirection=false&recordsPerPage=_100&showLotsInfoHidden=false&published=on&ur=on&customerPlace=5277335%2C5277327&customerPlaceCodes=%2C&updateDateFrom={frm}&updateDateTo={to}'
    print(url)
    return url

def request_url(url):
    """Получаем страницу"""
    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) '
                             'AppleWebKit/537.36 (KHTML, like Gecko) '
                             'Chrome/71.0.3578.98 Safari/537.36'}
    try:
        r = requests.get(url, headers=headers).text
        soup = bs(r, 'html.parser')
        return soup
    except Exception as e:
        print(e)

def new_day_xlsx(listik):
    wb = openpyxl.load_workbook(record_file)
    sheet = wb["1"]
    for row in range(len(listik)):
        for col in range(len(listik[row])):
            sheet.cell(row=row+2, column=col+1).value = listik[row][col]
    wb.save(filename=record_file)
    wb.close()

isha = 0
def parser_start(soup, base_of_post):
    """Парсер данных"""
    listik = []
    isha = 0
    block = soup.find_all(class_="search-registry-entry-block box-shadow-search-input")   #every zakupka
    print(len(block))
    try:
        for item in block:
            text_np = text_post_inn = text_url_kontr = text_tp = text_mp = text_ap = text_nomer = text_zakazchik = text_zakazchik_sokr = text_zakazchik_inn = ""
            text_ur = "https://zakupki.gov.ru/" + item.find('div', {'class': 'registry-entry__header-mid__number'}).findAll('a')[0].get(
                'href')  # url kontr                                                                  
            print(text_ur)
            nsoup = request_url(text_ur)
            # print(nsoup)
            poow = nsoup.find_all(class_="blockInfo__section section")

            try:
                for bss in poow:
                    if bss.find(class_="section__title").text == "Реестровый номер контракта":
                        text_url_kontr = bss.find('a').get('href')                          #краткое наименование заказчика
                if not text_url_kontr:
                    continue
                relink = request_url(text_url_kontr)
                text_pr = relink.find(class_='price').find(class_='cardMainInfo__content cost').text.strip()         #price
                dates = relink.find(class_="date mt-auto")
                text_date_zak_kon = dates.find(class_="cardMainInfo__content").text.strip()
                text_date_srok_isp = dates.find_all(class_="cardMainInfo__content")[1].text.strip()
                text_date_razm = dates.find_all(class_="cardMainInfo__content")[2].text.strip()
                text_date_obn = dates.find_all(class_="cardMainInfo__content")[3].text.strip()
                # print(text_date_zak_kon, text_date_srok_isp, text_date_razm, text_date_obn)

                blocks = relink.find_all(class_="blockInfo__section section")
                
                try:
                    for bss in blocks:
                        # print(bss.find(class_="section__title"))
                        if bss.find(class_="section__title").text == "Реестровый номер контракта":
                            # print(bss.find(class_="section__info").text)
                            text_nomer = bss.find(class_="section__info").text
                            continue
                        elif bss.find(class_="section__title").text.strip() == "Полное наименование заказчика":
                            # print(bss.find(class_="section__info").text)
                            text_zakazchik = bss.find(class_="section__info").text.strip()
                            continue
                        elif bss.find(class_="section__title").text.strip() == "Сокращенное наименование заказчика":
                            # print(bss.find(class_="section__info").text)
                            text_zakazchik_sokr = bss.find(class_="section__info").text.strip()
                            continue
                        elif bss.find(class_="section__title").text.strip() == "ИНН":
                            # print(bss.find(class_="section__info").text)
                            text_zakazchik_inn = bss.find(class_="section__info").text.strip()
                            continue
                except Exception as e:
                    print(e)
                # print(text_nomer, text_zakazchik, text_zakazchik_sokr, text_zakazchik_inn)

                np = relink.find(class_='tableBlock__col tableBlock__col_first text-break').text.split('\n')
                post_inn = relink.find(class_='tableBlock__col tableBlock__col_first text-break').find_all(class_="section")#[1].text.strip().split('\n')[1]
                for h in post_inn:
                    if h.text.strip().split('\n')[0] == "ИНН:":
                        text_post_inn = h.text.strip().split('\n')[1]
                # print(text_post_inn)
                for i in np:
                    if i != "":
                        text_np = i.strip()     #название поставщика
                        try:
                            text_np_ooo = text_np[1+ text_np.find("("):text_np.find(")")].strip()
                            ip = text_np_ooo.find("Индивидуальный")
                            if (ip != -1):
                                text_np_ooo = "ИП " + text_np_ooo[:ip]
                            else:
                                iooo = text_np_ooo.upper().find("ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ")
                                if iooo != -1:
                                    text_np_ooo = text_np_ooo[:iooo] + "ООО " + text_np_ooo[iooo + 41:]
                                ipao = text_np_ooo.upper().find("ПУБЛИЧНОЕ АКЦИОНЕРНОЕ ОБЩЕСТВО")
                                if ipao != -1:
                                    text_np_ooo = text_np_ooo[:ipao] + "ПАО " + text_np_ooo[ipao + 41:]
                                izao = text_np_ooo.upper().find("ЗАКРЫТОЕ АКЦИОНЕРНОЕ ОБЩЕСТВО")
                                if izao != -1:
                                    text_np_ooo = text_np_ooo[:izao] + "ЗАО " + text_np_ooo[izao + 31:]
                                ia = text_np_ooo.upper().find("АКЦИОНЕРНОЕ ОБЩЕСТВО")
                                if ia != -1:
                                    text_np_ooo = text_np_ooo[:ia] + "АО " + text_np_ooo[ia + 21:]
                                iz = text_np_ooo.upper().find("ЗАКРЫТОЕ АКЦИОНЕРНОЕ ОБЩЕСТВО")
                                if iz != -1:
                                    text_np_ooo = text_np_ooo[:iz] + "ЗАО " + text_np_ooo[iz + 30:]
                                iano = text_np_ooo.upper().find("АВТОНОМНАЯ НЕКОММЕРЧЕСКАЯ ОРГАНИЗАЦИЯ")
                                if iano != -1:
                                    text_np_ooo = text_np_ooo[:iano] + "АНО " + text_np_ooo[iano + 38:]
                                ifgao = text_np_ooo.upper().find("ФЕДЕРАЛЬНОЕ ГОСУДАРСТВЕННОЕ УНИТАРНОЕ ПРЕДПРИЯТИЕ")
                                if ifgao != -1:
                                    text_np_ooo = text_np_ooo[:ifgao] + "ФГУП " + text_np_ooo[ifgao + 38:]
                        except Exception as e:
                            print("ooo rip", e)
                        break
                
                

                col = relink.find_all(class_="col")[-1]  
                # print(col)
                postav = col.find_all(class_="tableBlock__col")
                length = len(postav)
                if (length == 15):
                    length = 10
                i_adres_mest = i_tel_po = 0
                for i in range(length):
                    # print(postav[i].text.replace('\n', '').strip())
                    if postav[i].text.replace('\n', '').strip() == "Адрес места нахождения":
                        i_adres_mest = i
                    elif postav[i].text.replace('\n', '').strip() == "Телефон, электронная почта":
                        i_tel_po = i
                    
                # print(i_adres_mest, i_tel_po)
                if i_adres_mest:
                    text_ap = postav[length // 2 + i_adres_mest].text.replace('\n', '').strip()
                else:
                    text_ap = ""
                if i_tel_po:
                    tp = postav[length // 2 + i_tel_po].text.replace('\n', '').strip()
                    ii = 0
                    for ii, c in enumerate(tp):
                        if (c.isalpha() or c.isnumeric()) and ii > 21:
                            break
                    if ii == 0:
                        text_tp = tp.strip()
                        text_mp = ""
                    else:
                        text_tp = tp[:ii].strip()   #телефон
                        text_mp = tp[ii:].strip()    #мыло   
                else:
                    text_tp = text_mp = ""

            except Exception as ex:
                1 + 1
            flag = 0
            for x in listik:
                if text_np in x:
                    flag = 1
                    break
            if flag ==1 or text_np in base_of_post:
                print("Данный поставщик уже присутсвует в базе", text_ur, text_np)
                continue
            data_item = (text_ur, text_url_kontr, text_nomer, text_pr, text_date_zak_kon, text_date_srok_isp, text_date_razm, text_date_obn, text_zakazchik, text_zakazchik_sokr, text_zakazchik_inn, text_np, text_np_ooo, text_post_inn, text_ap, text_tp, text_mp)
            # print(data_item)
            listik.append(data_item)
    except Exception as e:
        print(e)
    if (len(listik) > 0):
        new_day_xlsx(listik)


def return_column_from_excel():
    l = []
    wb = openpyxl.load_workbook(base_file)
    sheet = wb['1']
    for rowOfCellObjects in sheet['L1':'L999']:
        for cellObj in rowOfCellObjects:
            # print(cellObj.coordinate, cellObj.value)
            # v = cellObj.value
            l.append(cellObj.value)
    return(l)

head = {
                    "0":["Ссылка на Решение об одностороннем отказе от исполнения контракта"],
                    "1":["Ссылка на контракт"],
                    "2":["Реестровый номер контракта"],
                   "3": ["Цена контракта"],
                   "4": ["Заключение контракта"],
                   "5": ["Срок исполнения"],
                  "6":  ["Размещен контракт в реестре контрактов"],
                  "7":  ["Обновлен контракт в реестре контрактов"],
                   "8": ["Полное наименование заказчика"],
                   "9": ["Сокращенное наименование заказчика"],
                   "10": ["ИНН заказчика"],
                    "11":["Наименование поставщика"],
                   "12": ["Сокращенное наименование поставщика"],
                   "13": ["ИНН поставщика"],
                   "14": ["Адрес поставщика"],
                   "15": ["Телефон поставщика"],
                   "16": ["Почта Поставщика"]
                }

def file_create():
    df = pd.DataFrame(head)
    df.to_excel(record_file, sheet_name="1", index=False, header=False)

def paste_to_base(base_file):
    dfs = pd.read_excel(record_file, sheet_name="1")
    if (len(dfs.index) < 1):
        return (-1)
    base = pd.read_excel(base_file, sheet_name="1")
    flag = pd.DataFrame({'Ссылка на Решение об одностороннем отказе от исполнения контракта':[to]})
    # startrow = writer.sheets['Sheet1'].max_row
    dfs = pd.read_excel(record_file, sheet_name="1")
    # print(dfs.to_dict())
    res = pd.concat([flag, dfs])
    res = base.append([flag, dfs])
    res.to_excel(base_file, sheet_name="1", index=False, header=False)
    return (1)

def main():
    print(f"from {frm}\nto {to}")
    print("New file: " + record_file)
    file_create()
    base_of_post = return_column_from_excel()
    print("\n\n", base_of_post, "\n\n")
    print ('Начинаю сбор данных, подождите...')
    ist = 1
    while (pages >= ist):
        arr_url = urrll(ist)
        ist = ist + 1
        date_array = request_url(arr_url)
        parser_start(date_array, base_of_post)
        break
    if (input(f"Put data to base? ") == "y"):
        print("putting to base...")
        chek = paste_to_base(base_file)
        if (chek == 1
            and input(f"Can I send {record_file} and move to history/ ?: ") == "y"):
                send_email(record_file)
                shutil.move(record_file, f"history/{record_file}")
        elif (chek == -1):
            print("No new contracts")
    else:
        return

if __name__=="__main__":
    main()
    